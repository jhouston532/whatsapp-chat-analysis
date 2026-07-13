#!/usr/bin/env python3
"""
word_count_by_sender.py

Parses a WhatsApp group chat export (.txt) and builds a spreadsheet of
word counts per speaker:

    Word    | Alice | Bob | Carl | ... | Total
    --------|-------|-----|------|-----|------
    the     |   3   |  4  |  2   | ... |  9
    crypto  |   0   |  1  |  5   | ... |  6

Usage:
    python word_count_by_sender.py chat.txt
    python word_count_by_sender.py chat.txt -o word_counts.xlsx
    python word_count_by_sender.py chat.txt --top-n 300
    python word_count_by_sender.py chat.txt --include-system --min-count 2
"""

import argparse
import re
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Chat parsing (same approach as mean_time_between_messages.py)
# ---------------------------------------------------------------------------

BIDI_CHARS = re.compile(
    "[\u200e\u200f\u202a-\u202e\u2066-\u2069\u061c\u202f]"
)

MESSAGE_RE = re.compile(
    r"^\[(\d{1,2}/\d{1,2}/\d{2,4}),\s*(\d{1,2}:\d{2}:\d{2}\s?[APap][Mm])\]\s*"
    r"(?P<sender>[^:]+):\s?(?P<message>.*)$"
)

SYSTEM_MSG_MARKERS = (
    "Messages and calls are end-to-end encrypted",
    "created this group",
    "added you",
    "added ~",
    "changed the settings",
    "changed this group",
    "reset this group's invite link",
    "changed the group description",
    "<Media omitted>",
    "This message was deleted",
)

# Words too common to be interesting in most analyses; excluded by default.
DEFAULT_STOPWORDS = {
    "the", "a", "an", "and", "or", "but", "is", "are", "was", "were", "be",
    "been", "being", "to", "of", "in", "on", "at", "for", "with", "as",
    "it", "this", "that", "i", "you", "he", "she", "we", "they", "my",
    "your", "his", "her", "our", "their", "me", "him", "them", "us",
    "do", "does", "did", "have", "has", "had", "will", "would", "can",
    "could", "should", "not", "so", "if", "then", "there", "just",
}


def clean_line(line: str) -> str:
    return BIDI_CHARS.sub(" ", line).strip()


def is_system_message(message: str) -> bool:
    return any(marker in message for marker in SYSTEM_MSG_MARKERS)


def parse_chat(path: str):
    """Yield {'sender': ..., 'message': ...} dicts for every message."""
    entries = []
    current = None

    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = clean_line(raw_line)
            if not line:
                continue

            match = MESSAGE_RE.match(line)
            if match:
                _, _, sender, message = match.groups()
                current = {"sender": sender.strip(), "message": message}
                entries.append(current)
            else:
                if current is not None:
                    current["message"] += "\n" + line

    return entries


# ---------------------------------------------------------------------------
# Word counting
# ---------------------------------------------------------------------------

WORD_RE = re.compile(r"[A-Za-z']+")


def tokenize(text: str):
    for raw in WORD_RE.findall(text.lower()):
        word = raw.strip("'")
        if word:
            yield word


def build_counts(entries, exclude_system, stopwords):
    """Return (word_counts: {word: {sender: count}}, senders: sorted list)."""
    word_counts = defaultdict(lambda: defaultdict(int))
    senders = set()

    for entry in entries:
        if exclude_system and is_system_message(entry["message"]):
            continue
        sender = entry["sender"]
        senders.add(sender)
        for word in tokenize(entry["message"]):
            if word in stopwords:
                continue
            word_counts[word][sender] += 1

    return word_counts, sorted(senders, key=str.lower)


# ---------------------------------------------------------------------------
# Spreadsheet output
# ---------------------------------------------------------------------------

def write_spreadsheet(word_counts, senders, out_path, top_n=None, min_count=1):
    totals = {word: sum(counts.values()) for word, counts in word_counts.items()}

    rows = [w for w, t in totals.items() if t >= min_count]
    rows.sort(key=lambda w: totals[w], reverse=True)
    if top_n:
        rows = rows[:top_n]

    wb = Workbook()
    ws = wb.active
    ws.title = "Word Counts"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name="Arial")

    headers = ["Word"] + senders + ["Total"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, word in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=word).font = body_font
        counts = word_counts[word]
        for c, sender in enumerate(senders, start=2):
            cell = ws.cell(row=r, column=c, value=counts.get(sender, 0))
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center")
        total_cell = ws.cell(row=r, column=len(headers), value=totals[word])
        total_cell.font = Font(name="Arial", bold=True)
        total_cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws.freeze_panes = "B2"

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(
        description="Count word frequency per sender in a WhatsApp export and write an xlsx."
    )
    parser.add_argument("chat_file", help="Path to the exported WhatsApp .txt file")
    parser.add_argument(
        "-o", "--output", default="word_counts.xlsx",
        help="Output .xlsx path (default: word_counts.xlsx)",
    )
    parser.add_argument(
        "--top-n", type=int, default=None,
        help="Only keep the N most frequent words overall (default: all)",
    )
    parser.add_argument(
        "--min-count", type=int, default=1,
        help="Drop words with a total count below this threshold (default: 1)",
    )
    parser.add_argument(
        "--include-system", action="store_true",
        help="Include WhatsApp system messages (joins, settings changes, etc.)",
    )
    parser.add_argument(
        "--no-stopwords", action="store_true",
        help="Don't filter out common stopwords (the, a, is, ...)",
    )
    args = parser.parse_args()

    entries = parse_chat(args.chat_file)
    if not entries:
        print("No messages could be parsed from this file.")
        return

    stopwords = set() if args.no_stopwords else DEFAULT_STOPWORDS
    word_counts, senders = build_counts(
        entries, exclude_system=not args.include_system, stopwords=stopwords
    )

    if not word_counts:
        print("No words found after filtering.")
        return

    write_spreadsheet(
        word_counts, senders, args.output, top_n=args.top_n, min_count=args.min_count
    )

    print(f"Parsed {len(entries)} messages from {len(senders)} senders.")
    print(f"Wrote {len(word_counts)} unique words to {args.output}"
          + (f" (top {args.top_n})" if args.top_n else ""))


if __name__ == "__main__":
    main()