#!/usr/bin/env python3
"""
activity_heatmap_table.py

Parses a WhatsApp group chat export (.txt) and builds a sender x time-bin
activity table (message counts per sender per time bucket), written to xlsx:

    Time              | Sender1 | Sender2 | ...
    ------------------|---------|---------|----
    2026-05-05 09:30  |    2    |    0    | ...
    2026-05-05 10:00  |    0    |    3    | ...

Usage:
    python activity_heatmap_table.py chat.txt
    python activity_heatmap_table.py chat.txt -o activity.xlsx --bin-minutes 60
    python activity_heatmap_table.py chat.txt --sort total --include-system
"""

import argparse
import re
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ---------------------------------------------------------------------------
# Chat parsing (same approach as the other chat-analysis scripts)
# ---------------------------------------------------------------------------

BIDI_CHARS = re.compile(
    "[\u200e\u200f\u202a-\u202e\u2066-\u2069\u061c\u202f]"
)

MESSAGE_RE = re.compile(
    r"^\[(\d{1,2}/\d{1,2}/\d{2,4}),\s*(\d{1,2}:\d{2}:\d{2}\s?[APap][Mm])\]\s*"
    r"(?P<sender>[^:]+):\s?(?P<message>.*)$"
)

SYSTEM_MSG_MARKERS = (
    "Messages and calls are end-to-end encrypted",
    "created this group",
    "added you",
    "added ~",
    "changed the settings",
    "changed this group",
    "reset this group's invite link",
    "changed the group description",
    "<Media omitted>",
    "This message was deleted",
)


def clean_line(line: str) -> str:
    return BIDI_CHARS.sub(" ", line).strip()


def is_system_message(message: str) -> bool:
    return any(marker in message for marker in SYSTEM_MSG_MARKERS)


def parse_datetime(date_str: str, time_str: str):
    time_str = time_str.replace(" ", "")
    candidates = [
        "%m/%d/%y,%I:%M:%S%p",
        "%m/%d/%Y,%I:%M:%S%p",
        "%d/%m/%y,%I:%M:%S%p",
        "%d/%m/%Y,%I:%M:%S%p",
    ]
    combined = f"{date_str},{time_str}"
    for fmt in candidates:
        try:
            return datetime.strptime(combined, fmt)
        except ValueError:
            continue
    return None


def parse_chat(path: str):
    """Yield {'dt': datetime, 'sender': ..., 'message': ...} for every message."""
    entries = []
    current = None

    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = clean_line(raw_line)
            if not line:
                continue

            match = MESSAGE_RE.match(line)
            if match:
                date_str, time_str, sender, message = match.groups()
                dt = parse_datetime(date_str, time_str)
                if dt is None:
                    if current is not None:
                        current["message"] += "\n" + line
                    continue
                current = {"dt": dt, "sender": sender.strip(), "message": message}
                entries.append(current)
            else:
                if current is not None:
                    current["message"] += "\n" + line

    return entries


# ---------------------------------------------------------------------------
# Binning
# ---------------------------------------------------------------------------

def floor_to_bin(dt: datetime, bin_minutes: int) -> datetime:
    """Round a timestamp down to the start of its bin."""
    minutes_since_midnight = dt.hour * 60 + dt.minute
    floored = (minutes_since_midnight // bin_minutes) * bin_minutes
    return dt.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=floored)


def build_activity_table(entries, bin_minutes, exclude_system):
    if exclude_system:
        entries = [e for e in entries if not is_system_message(e["message"])]
    if not entries:
        return {}, [], []

    entries.sort(key=lambda e: e["dt"])
    first_bin = floor_to_bin(entries[0]["dt"], bin_minutes)
    last_bin = floor_to_bin(entries[-1]["dt"], bin_minutes)

    bins = []
    cursor = first_bin
    step = timedelta(minutes=bin_minutes)
    while cursor <= last_bin:
        bins.append(cursor)
        cursor += step

    counts = defaultdict(lambda: defaultdict(int))
    first_seen = {}
    for e in entries:
        sender = e["sender"]
        b = floor_to_bin(e["dt"], bin_minutes)
        counts[sender][b] += 1
        if sender not in first_seen:
            first_seen[sender] = e["dt"]

    senders = list(counts.keys())
    return counts, senders, bins, first_seen


# ---------------------------------------------------------------------------
# Spreadsheet output
# ---------------------------------------------------------------------------

def write_spreadsheet(counts, senders, bins, out_path, sort_mode, first_seen):
    if sort_mode == "alpha":
        senders_sorted = sorted(senders, key=str.lower)
    elif sort_mode == "total":
        senders_sorted = sorted(
            senders, key=lambda s: sum(counts[s].values()), reverse=True
        )
    else:  # first-seen (default) — matches chronological appearance in the chat
        senders_sorted = sorted(senders, key=lambda s: first_seen[s])

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name="Arial")
    label_font = Font(name="Arial", bold=True)

    # Header row: "time" then one column per bin
    ws.cell(row=1, column=1, value="time").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for c, b in enumerate(bins, start=2):
        cell = ws.cell(row=1, column=c, value=b.strftime("%Y-%m-%d %H:%M"))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Force text format so senders like "+1 (626) 512-8920" and these
        # timestamp strings are never mis-parsed as formulas/numbers when
        # opened in Sheets/Excel.
        cell.number_format = "@"

    # Data rows: one per sender, force sender names to text format too
    for r, sender in enumerate(senders_sorted, start=2):
        name_cell = ws.cell(row=r, column=1, value=sender)
        name_cell.font = label_font
        name_cell.number_format = "@"
        for c, b in enumerate(bins, start=2):
            value = counts[sender].get(b, 0)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 24
    for c in range(2, len(bins) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "B2"

    # Built-in heatmap: 3-color scale over the numeric grid
    if bins and senders_sorted:
        last_row = len(senders_sorted) + 1
        last_col = len(bins) + 1
        data_range = f"B2:{get_column_letter(last_col)}{last_row}"
        rule = ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws.conditional_formatting.add(data_range, rule)

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(
        description="Build a sender x time-bin activity table from a WhatsApp export."
    )
    parser.add_argument("chat_file", help="Path to the exported WhatsApp .txt file")
    parser.add_argument(
        "-o", "--output", default="activity_table.xlsx",
        help="Output .xlsx path (default: activity_table.xlsx)",
    )
    parser.add_argument(
        "--bin-minutes", type=int, default=60,
        help="Size of each time bucket in minutes (default: 60, i.e. 1-hour segments)",
    )
    parser.add_argument(
        "--sort", choices=["first-seen", "total", "alpha"], default="first-seen",
        help="Order of sender rows: first-seen (default, chronological), "
        "total (most active first), or alpha",
    )
    parser.add_argument(
        "--include-system", action="store_true",
        help="Include WhatsApp system messages (joins, settings changes, etc.)",
    )
    args = parser.parse_args()

    entries = parse_chat(args.chat_file)
    if not entries:
        print("No messages could be parsed from this file.")
        return

    counts, senders, bins, first_seen = build_activity_table(
        entries, args.bin_minutes, exclude_system=not args.include_system
    )
    if not bins:
        print("No messages left after filtering.")
        return

    write_spreadsheet(counts, senders, bins, args.output, args.sort, first_seen)

    print(f"Parsed {len(entries)} messages.")
    print(f"{len(senders)} senders x {len(bins)} time bins "
          f"({args.bin_minutes}-minute segments) written to {args.output}")


if __name__ == "__main__":
    main()